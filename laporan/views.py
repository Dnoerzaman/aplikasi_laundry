import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views.generic import FormView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin

from .forms import ReportFilterForm
from checklist.models import PenerimaanLinen, BeratLinenHarian
from chemical.models import PemakaianChemical, StokChemical
from stok_linen.models import TransaksiLinen
from aset_laundry.models import Aset
from schedule.models import Tugas

class ReportDashboardView(LoginRequiredMixin, FormView):
    template_name = 'laporan/report_dashboard.html'
    form_class = ReportFilterForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Rekap Laporan'
        return context

class BaseExportView(LoginRequiredMixin, View):
    """
    Base View untuk menangani logika ekspor Excel dengan filter tanggal.
    """
    def get(self, request, *args, **kwargs):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if not all([start_date, end_date]):
            return HttpResponse("Tanggal Mulai dan Tanggal Selesai harus diisi.", status=400)
        
        # Asumsikan field tanggal bernama 'tanggal'
        queryset = self.get_queryset().filter(tanggal__range=[start_date, end_date])

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.get_sheet_title()

        headers = self.get_headers()
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 22

        for row_num, data_object in enumerate(queryset, 2):
            row_data = self.get_row_data(data_object)
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename()}"'
        workbook.save(response)
        return response

    def get_queryset(self):
        raise NotImplementedError
    def get_headers(self):
        raise NotImplementedError
    def get_row_data(self, data_object):
        raise NotImplementedError
    def get_sheet_title(self):
        return "Laporan"
    def get_filename(self):
        return "laporan.xlsx"

class BaseSnapshotExportView(BaseExportView):
    """
    Base View untuk laporan snapshot yang tidak memerlukan filter tanggal.
    """
    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.get_sheet_title()
        headers = self.get_headers()
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25
        for row_num, data_object in enumerate(queryset, 2):
            row_data = self.get_row_data(data_object)
            for col_num, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename()}"'
        workbook.save(response)
        return response

# --- EXPORT VIEWS ---

class ExportChecklistExcel(BaseExportView):
    def get_queryset(self):
        return PenerimaanLinen.objects.all().prefetch_related('items')

    def get(self, request, *args, **kwargs):
        # Override method 'get' karena strukturnya berbeda (nested)
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if not all([start_date, end_date]):
            return HttpResponse("Tanggal Mulai dan Tanggal Selesai harus diisi.", status=400)
        
        queryset = self.get_queryset().filter(tanggal__range=[start_date, end_date])
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Penerimaan Linen"
        headers = ["Tanggal", "Jam", "Ruangan", "Petugas", "Nama Item", "Jumlah", "Kondisi", "Keterangan Item"]
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            worksheet.column_dimensions[get_column_letter(col_num)].width = 22
        
        row_num = 2
        for penerimaan in queryset:
            for item in penerimaan.items.all():
                row_data = [
                    penerimaan.tanggal, penerimaan.jam, penerimaan.get_ruangan_display(),
                    penerimaan.petugas.username, item.get_nama_item_display(),
                    item.jumlah, item.get_kondisi_display(), item.keterangan
                ]
                for col_num, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num).value = cell_value
                row_num += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="laporan_penerimaan_linen.xlsx"'
        workbook.save(response)
        return response

class ExportPemakaianChemicalExcel(BaseExportView):
    def get_queryset(self):
        return PemakaianChemical.objects.all()
    def get_headers(self):
        return ["Tanggal", "Nama Chemical", "Jumlah", "Satuan", "Petugas", "Keterangan"]
    def get_row_data(self, data_object):
        return [
            data_object.tanggal, data_object.chemical.get_nama_chemical_display(),
            data_object.jumlah, data_object.chemical.get_unit_display(),
            data_object.petugas.username, data_object.keterangan
        ]
    def get_sheet_title(self):
        return "Pemakaian Chemical"
    def get_filename(self):
        return "laporan_pemakaian_chemical.xlsx"

class ExportTransaksiLinenExcel(BaseExportView):
    def get_queryset(self):
        return TransaksiLinen.objects.all()
    def get_headers(self):
        return ["Tanggal", "Nama Linen", "Ruangan", "Jenis Transaksi", "Jumlah", "Petugas", "Keterangan"]
    def get_row_data(self, data_object):
        return [
            data_object.tanggal, data_object.stok_linen.get_nama_linen_display(),
            data_object.stok_linen.get_ruangan_display(), data_object.get_jenis_transaksi_display(),
            data_object.jumlah, data_object.petugas.username, data_object.keterangan
        ]
    def get_sheet_title(self):
        return "Transaksi Linen"
    def get_filename(self):
        return "laporan_transaksi_linen.xlsx"

class ExportBeratLinenExcel(BaseExportView):
    def get_queryset(self):
        return BeratLinenHarian.objects.all()
    def get_headers(self):
        return ["Tanggal", "Ruangan", "Shift", "Total Berat (Kg)", "Petugas"]
    def get_row_data(self, data_object):
        return [
            data_object.tanggal, data_object.get_ruangan_display(),
            data_object.get_shift_display(), data_object.total_berat,
            data_object.petugas.username,
        ]
    def get_sheet_title(self):
        return "Berat Linen Harian"
    def get_filename(self):
        return "laporan_berat_linen_harian.xlsx"

class ExportAsetExcel(BaseSnapshotExportView):
    def get_queryset(self):
        return Aset.objects.all()
    def get_headers(self):
        return ["Nama Barang", "Jumlah", "Satuan", "Merk/Tipe", "SN", "Tahun Pengadaan", "Tanggal Input", "Keterangan"]
    def get_row_data(self, data_object):
        return [
            data_object.nama_barang, data_object.jumlah, data_object.get_satuan_display(),
            data_object.merk, data_object.serial_number, data_object.tahun_pengadaan,
            data_object.tanggal_input, data_object.keterangan
        ]
    def get_sheet_title(self):
        return "Daftar Aset"
    def get_filename(self):
        return "laporan_daftar_aset.xlsx"

class ExportScheduleExcel(BaseSnapshotExportView):
    def get_queryset(self):
        return Tugas.objects.all()
    def get_headers(self):
        return ["Judul Tugas", "Status", "Penanggung Jawab", "Target Waktu", "Bulan & Tahun", "Deskripsi"]
    def get_row_data(self, data_object):
        return [
            data_object.judul, data_object.get_status_display(),
            data_object.penanggung_jawab.username if data_object.penanggung_jawab else '-',
            data_object.get_target_waktu_display(), data_object.periode, data_object.deskripsi
        ]
    def get_sheet_title(self):
        return "Rencana Kerja"
    def get_filename(self):
        return "laporan_rencana_kerja.xlsx"

class ExportStokChemicalExcel(BaseSnapshotExportView):
    def get_queryset(self):
        return StokChemical.objects.all()
    def get_headers(self):
        return ["Nama Chemical", "Jumlah Stok", "Satuan", "Update Terakhir"]
    def get_row_data(self, data_object):
        return [
            data_object.get_nama_chemical_display(), data_object.jumlah_stok,
            data_object.get_unit_display(), data_object.update_terakhir.strftime("%Y-%m-%d")
        ]
    def get_sheet_title(self):
        return "Stok Chemical"
    def get_filename(self):
        return "laporan_stok_chemical.xlsx"
